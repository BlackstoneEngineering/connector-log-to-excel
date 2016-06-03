from openpyxl import Workbook, load_workbook
import mbed_connector_api as mc
import time
import datetime
API_KEY = 'Copy from connector.mbed.com/'
endpointName = 'ChangeMe'
resourceName = 'ChangeMe'

def callbackFn(data):
	log = [datetime.datetime.now().isoformat(), data.result]
	print(log) #log result to screen
	try:
		wb = load_workbook('results.xlsx')
	except:
		wb = Workbook()
	ws = wb.active
	ws.append(log)
	wb.save('results.xlsx')

def main():
	# setup the logbook

	# setup connector stuff
	x = mc.connector(API_KEY)
	x.startLongPolling()

	while(1):
		# every 5 seconds read the value and record it
		x.getResourceValue(endpointName,resourceName,callbackFn)
		time.sleep(5)

if __name__ == '__main__':
	main()